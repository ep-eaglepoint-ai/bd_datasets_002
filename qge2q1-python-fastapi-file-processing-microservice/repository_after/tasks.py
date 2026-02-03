from __future__ import annotations

import asyncio
import csv
import math
import re
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from celery.exceptions import SoftTimeLimitExceeded
from openpyxl import load_workbook
from redis.asyncio import Redis
from sqlalchemy import insert, select, update
from sqlalchemy.ext.asyncio import AsyncSession

from .celery_app import celery_app
from .config import settings
from .db import create_engine, create_sessionmaker
from .models import Job, JobStatus, LoadedRow, ProcessingError
from .storage import job_storage_path
from .webhook import deliver_webhook


def _now() -> datetime:
    return datetime.now(timezone.utc)


def _is_null(value: Any) -> bool:
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except Exception:
        pass
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


_INT_RE = re.compile(r"^[+-]?\d+$")
_FLOAT_RE = re.compile(r"^[+-]?(?:\d+\.\d*|\d*\.\d+)(?:[eE][+-]?\d+)?$|^[+-]?\d+(?:[eE][+-]?\d+)$")


def _transform_value(value: Any) -> Any:
    if _is_null(value):
        return None

    if isinstance(value, (int, bool)):
        return int(value) if isinstance(value, bool) else value

    if isinstance(value, float):
        if not math.isfinite(value):
            return None
        if value.is_integer():
            # avoid storing 1.0 as float
            as_int = int(value)
            return as_int
        return value

    if isinstance(value, datetime):
        # JSON-friendly
        return value.astimezone(timezone.utc).isoformat()

    if isinstance(value, str):
        s = value.strip()
        if s == "":
            return None

        # Basic numeric parsing for "data type" validation.
        if _INT_RE.match(s):
            try:
                return int(s)
            except Exception:
                return s

        if _FLOAT_RE.match(s):
            try:
                f = float(s)
                return f if math.isfinite(f) else None
            except Exception:
                return s

        return s

    # Fallback: stringify unknown types
    try:
        return str(value)
    except Exception:
        return None


def _transform_row(row: dict[str, Any]) -> dict[str, Any]:
    return {str(k): _transform_value(v) for k, v in row.items()}


def _validate_row(row: dict[str, Any]) -> list[tuple[str | None, str, str, str | None]]:
    errors: list[tuple[str | None, str, str, str | None]] = []
    for col, val in row.items():
        if _is_null(val):
            errors.append((col, "NULL", "Value is required", None if val is None else str(val)))
        else:
            if isinstance(val, str) and len(val) > 2000:
                errors.append((col, "CONSTRAINT", "Value too long", val[:2000]))
            # Basic data type / constraint validation
            if isinstance(val, int):
                if val < -(2**63) or val > (2**63 - 1):
                    errors.append((col, "CONSTRAINT", "Integer out of range", str(val)))
            elif isinstance(val, float):
                if not math.isfinite(val):
                    errors.append((col, "TYPE", "Non-finite float", str(val)))
            elif isinstance(val, (str, bool)):
                # allowed
                pass
            else:
                errors.append((col, "TYPE", "Unsupported value type", str(type(val))))
    return errors


def _should_persist(last_persist: float, now: float, interval: float) -> bool:
    return (now - last_persist) >= interval


async def _get_job(session: AsyncSession, job_id: uuid.UUID) -> Job | None:
    res = await session.execute(select(Job).where(Job.id == job_id))
    return res.scalar_one_or_none()


async def _update_job_progress(
    session: AsyncSession,
    job_id: uuid.UUID,
    *,
    status: JobStatus | None = None,
    progress: int | None = None,
    rows_processed: int | None = None,
    rows_failed: int | None = None,
    error_message: str | None = None,
    started_at: datetime | None = None,
    completed_at: datetime | None = None,
) -> None:
    values: dict[str, Any] = {}
    if status is not None:
        values["status"] = status
    if progress is not None:
        values["progress"] = max(0, min(100, int(progress)))
    if rows_processed is not None:
        values["rows_processed"] = int(rows_processed)
    if rows_failed is not None:
        values["rows_failed"] = int(rows_failed)
    if error_message is not None:
        values["error_message"] = error_message
    if started_at is not None:
        values["started_at"] = started_at
    if completed_at is not None:
        values["completed_at"] = completed_at

    if not values:
        return

    await session.execute(update(Job).where(Job.id == job_id).values(**values))
    await session.commit()


async def _log_errors(
    session: AsyncSession,
    job_id: uuid.UUID,
    row_number: int,
    errors: list[tuple[str | None, str, str, str | None]],
) -> None:
    payload = []
    for (column_name, error_type, error_message, raw_value) in errors:
        payload.append(
            {
                "id": uuid.uuid4(),
                "job_id": job_id,
                "row_number": int(row_number),
                "column_name": column_name,
                "error_type": error_type,
                "error_message": error_message,
                "raw_value": raw_value,
            }
        )
    if payload:
        await session.execute(insert(ProcessingError), payload)


async def _insert_loaded_rows(
    session: AsyncSession,
    *,
    job_id: uuid.UUID,
    rows: list[tuple[int, dict[str, Any]]],
) -> None:
    if not rows:
        return
    payload = []
    for (row_number, data) in rows:
        payload.append(
            {
                "id": uuid.uuid4(),
                "job_id": job_id,
                "row_number": int(row_number),
                "data": data,
            }
        )
    await session.execute(insert(LoadedRow), payload)


async def _process_csv(path: Path, job_id: uuid.UUID, session: AsyncSession) -> tuple[int, int]:
    # Determine total rows (excluding header) for progress.
    total_rows = 0
    with path.open("r", newline="", encoding="utf-8", errors="ignore") as f:
        reader = csv.reader(f)
        try:
            next(reader)
        except StopIteration:
            total_rows = 0
        else:
            total_rows = sum(1 for _ in reader)

    rows_seen = 0
    rows_failed = 0

    last_persist = time.monotonic()

    pending_loaded: list[tuple[int, dict[str, Any]]] = []
    pending_error_inserts = 0

    for chunk in pd.read_csv(path, chunksize=10_000):
        # cancellation check before chunk
        job = await _get_job(session, job_id)
        if not job or job.status == JobStatus.CANCELLED:
            break

        for _, row in chunk.iterrows():
            rows_seen += 1
            raw = {str(k): row[k] for k in row.index}
            transformed = _transform_row(raw)
            errs = _validate_row(transformed)
            if errs:
                rows_failed += 1
                await _log_errors(session, job_id, rows_seen, errs)
                pending_error_inserts += len(errs)
            else:
                pending_loaded.append((rows_seen, transformed))

            now = time.monotonic()
            if _should_persist(last_persist, now, settings.progress_update_interval_seconds):
                last_persist = now

                # Persist loaded rows + errors and job progress at least every 5 seconds.
                await _insert_loaded_rows(session, job_id=job_id, rows=pending_loaded)
                pending_loaded.clear()

                progress = 100 if total_rows == 0 else int((rows_seen / max(total_rows, 1)) * 100)
                await _update_job_progress(
                    session,
                    job_id,
                    progress=progress,
                    rows_processed=rows_seen,
                    rows_failed=rows_failed,
                )

        # End of chunk: persist any pending loaded rows so we don't build up memory.
        if pending_loaded:
            await _insert_loaded_rows(session, job_id=job_id, rows=pending_loaded)
            pending_loaded.clear()

        # Ensure progress isn't stale for long-running chunks.
        now = time.monotonic()
        if _should_persist(last_persist, now, settings.progress_update_interval_seconds):
            last_persist = now
            progress = 100 if total_rows == 0 else int((rows_seen / max(total_rows, 1)) * 100)
            await _update_job_progress(
                session,
                job_id,
                progress=progress,
                rows_processed=rows_seen,
                rows_failed=rows_failed,
            )

    # final persist
    if pending_loaded:
        await _insert_loaded_rows(session, job_id=job_id, rows=pending_loaded)
        pending_loaded.clear()
    progress = 100 if total_rows == 0 else int((rows_seen / max(total_rows, 1)) * 100)
    await _update_job_progress(session, job_id, progress=progress, rows_processed=rows_seen, rows_failed=rows_failed)
    return rows_seen, rows_failed


async def _process_excel(path: Path, job_id: uuid.UUID, session: AsyncSession) -> tuple[int, int]:
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active

    total_rows = max(0, (ws.max_row or 0) - 1)

    rows_seen = 0
    rows_failed = 0

    header: list[str] = []
    last_persist = time.monotonic()

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_vals = next(rows_iter)
        header = [str(c) if c is not None else "" for c in header_vals]
    except StopIteration:
        header = []

    batch: list[dict[str, Any]] = []
    batch_row_numbers: list[int] = []

    for row_vals in rows_iter:
        rows_seen += 1

        # Check cancellation before building the next chunk.
        job = await _get_job(session, job_id)
        if not job or job.status == JobStatus.CANCELLED:
            break

        row_dict = {header[i] if i < len(header) else f"col_{i}": row_vals[i] for i in range(len(row_vals))}
        batch.append(_transform_row(row_dict))
        batch_row_numbers.append(rows_seen)

        if len(batch) >= 10_000:
            pending_loaded: list[tuple[int, dict[str, Any]]] = []
            for i, item in enumerate(batch):
                rn = batch_row_numbers[i]
                errs = _validate_row(item)
                if errs:
                    rows_failed += 1
                    await _log_errors(session, job_id, rn, errs)
                else:
                    pending_loaded.append((rn, item))

            await _insert_loaded_rows(session, job_id=job_id, rows=pending_loaded)
            batch.clear()
            batch_row_numbers.clear()

        now = time.monotonic()
        if _should_persist(last_persist, now, settings.progress_update_interval_seconds):
            last_persist = now
            progress = 100 if total_rows == 0 else int((rows_seen / max(total_rows, 1)) * 100)
            await _update_job_progress(session, job_id, progress=progress, rows_processed=rows_seen, rows_failed=rows_failed)

    if batch:
        pending_loaded: list[tuple[int, dict[str, Any]]] = []
        for i, item in enumerate(batch):
            rn = batch_row_numbers[i]
            errs = _validate_row(item)
            if errs:
                rows_failed += 1
                await _log_errors(session, job_id, rn, errs)
            else:
                pending_loaded.append((rn, item))
        await _insert_loaded_rows(session, job_id=job_id, rows=pending_loaded)

    progress = 100 if total_rows == 0 else int((rows_seen / max(total_rows, 1)) * 100)
    await _update_job_progress(session, job_id, progress=progress, rows_processed=rows_seen, rows_failed=rows_failed)

    wb.close()
    return rows_seen, rows_failed


async def _process_job_async(job_id_str: str) -> None:
    engine = create_engine()
    sessionmaker = create_sessionmaker(engine)

    job_id = uuid.UUID(job_id_str)

    async with sessionmaker() as session:
        job = await _get_job(session, job_id)
        if not job:
            return

        if job.status == JobStatus.CANCELLED:
            return

        await _update_job_progress(session, job_id, status=JobStatus.PROCESSING, started_at=_now(), error_message=None)

        path = job_storage_path(job_id_str, job.filename)

        try:
            if job.file_type == "csv":
                rows_processed, rows_failed = await _process_csv(path, job_id, session)
            elif job.file_type in {"xlsx", "xls"}:
                rows_processed, rows_failed = await _process_excel(path, job_id, session)
            else:
                raise ValueError("Unsupported file type")

            refreshed = await _get_job(session, job_id)
            if refreshed and refreshed.status == JobStatus.CANCELLED:
                await _update_job_progress(session, job_id, completed_at=_now())
                # cancelled jobs are retryable; webhook payload supports only COMPLETED|FAILED
                if job.webhook_url:
                    try:
                        await deliver_webhook(
                            job.webhook_url,
                            job_id=job_id_str,
                            status="FAILED",
                            rows_processed=rows_processed,
                            rows_failed=rows_failed,
                            completed_at=_now(),
                        )
                    except Exception:
                        pass
                return

            await _update_job_progress(session, job_id, status=JobStatus.COMPLETED, progress=100, completed_at=_now())

            if job.webhook_url:
                try:
                    await deliver_webhook(
                        job.webhook_url,
                        job_id=job_id_str,
                        status="COMPLETED",
                        rows_processed=rows_processed,
                        rows_failed=rows_failed,
                        completed_at=_now(),
                    )
                except Exception:
                    # Webhook failures should not fail the job processing.
                    pass

        except SoftTimeLimitExceeded:
            await _update_job_progress(
                session,
                job_id,
                status=JobStatus.FAILED,
                error_message="Worker soft time limit exceeded; job can be retried",
                completed_at=_now(),
            )
        except Exception as exc:  # noqa: BLE001
            await _update_job_progress(
                session,
                job_id,
                status=JobStatus.FAILED,
                error_message=str(exc),
                completed_at=_now(),
            )
            if job.webhook_url:
                try:
                    await deliver_webhook(
                        job.webhook_url,
                        job_id=job_id_str,
                        status="FAILED",
                        rows_processed=job.rows_processed,
                        rows_failed=job.rows_failed,
                        completed_at=_now(),
                    )
                except Exception:
                    pass

    await engine.dispose()


@celery_app.task(name="process_job")
def process_job(job_id: str) -> None:
    asyncio.run(_process_job_async(job_id))


async def check_redis_health(redis_url: str) -> bool:
    try:
        redis = Redis.from_url(redis_url)
        await redis.ping()
        await redis.aclose()
        return True
    except Exception:
        return False
